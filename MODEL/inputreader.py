import openpyxl as xl

def inp_mapper(omin, omax, nmin, nmax, data):
    orange = omax - omin
    nrange = nmax - nmin

    orange = max(1, orange)

    ndata = ((data - omin)*nrange/orange) + nmin
    return ndata

def cell_ind(v, i, r, C):
    if(v == 1):
        x = chr(i+65) + str(r+1)
        return x
    elif(v == 2):
        x = chr(65+C-1-i) + str(r+1)
        return x

def inparse(v, I, Af, filename):
    wb = xl.load_workbook('files\\'+ filename +'.xlsx')
    ws = wb.worksheets[0]

    R = ws.max_row
    C = ws.max_column

    val = [[0 for x in range(I)] for y in range(R)]
    max = [ws[cell_ind(v, i, 0, C)].value for i in range(I)]
    min = [ws[cell_ind(v, i, 0, C)].value for i in range(I)]

    nmax = 1
    nmin = 0

    for r in range(R):
        for i in range(I):
            x = cell_ind(v, i, r, C)
            val[r][i] = ws[x].value
            if(max[i] < val[r][i]):
                max[i] = val[r][i]
            if(min[i] > val[r][i]):
                min[i] = val[r][i]

    if(Af == 1 and v == 1):
        nmax = 1
        nmin = -1
    
    for r in range(R):
        for i in range(I):
            val[r][i] = inp_mapper(min[i], max[i], nmin, nmax, val[r][i])

    return val